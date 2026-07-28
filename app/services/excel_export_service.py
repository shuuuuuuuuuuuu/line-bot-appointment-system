"""Build Excel exports for admin."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _fmt_dt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    text = str(value).replace("T", " ")
    return text[:16] if len(text) >= 16 else text


def build_appointments_workbook(rows) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "預約明細"

    headers = [
        "編號",
        "客戶",
        "服務時間",
        "分類",
        "服務項目",
        "時長（分）",
        "備註",
        "匯款資訊",
        "金額",
        "狀態",
        "建立時間",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for row in rows:
        ws.append(
            [
                row.id,
                row.client_name,
                _fmt_dt(row.service_date_time),
                row.category_name or "",
                "、".join(row.service_names or []),
                row.total_duration if row.total_duration is not None else "",
                row.user_message or "",
                "已收到" if row.payment_proof_received else "尚未",
                row.total_price or 0,
                row.status_label,
                _fmt_dt(row.created_at),
            ]
        )

    widths = [8, 12, 18, 12, 28, 12, 36, 12, 10, 10, 18]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
